import streamlit as st
import pandas as pd
import requests
import FinanceDataReader as fdr

from io import StringIO, BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


st.set_page_config(
    page_title="단타 매매 분석기",
    page_icon="📈",
    layout="wide"
)

st.title("📈 단타 매매 후보 자동 분석기")
st.caption("네이버 금융 거래대금 상위 종목 기반 1차 필터링 웹앱")
st.warning("투자 참고용입니다. 매수·매도 판단은 본인 책임입니다.")

st.sidebar.header("필터 조건")

min_trade_value = st.sidebar.number_input("최소 거래대금(억)", value=500)
min_market_cap = st.sidebar.number_input("최소 시총(억)", value=3000)
max_market_cap = st.sidebar.number_input("최대 시총(억)", value=50000)
min_change = st.sidebar.number_input("최소 등락률(%)", value=5.0)
max_change = st.sidebar.number_input("최대 등락률(%)", value=20.0)
pages = st.sidebar.number_input("가져올 페이지 수", value=4, min_value=1, max_value=10)


headers = {
    "User-Agent": "Mozilla/5.0"
}


def parse_pct(val):
    s = str(val)
    s = s.replace("%", "")
    s = s.replace("+", "")
    s = s.replace(",", "")
    s = s.strip()

    try:
        return float(s)
    except:
        return 0


def get_naver_top_volume(sosok=0, pages=4):
    all_dfs = []

    for page in range(1, pages + 1):
        url = f"https://finance.naver.com/sise/sise_quant.naver?sosok={sosok}&page={page}"

        try:
            response = requests.get(url, headers=headers, timeout=10)
            response.encoding = "euc-kr"

            tables = pd.read_html(StringIO(response.text))

            df = tables[1]
            df = df.dropna(how="all").dropna(axis=1, how="all")
            df = df[df["종목명"].notna()]

            all_dfs.append(df)

        except Exception as e:
            st.error(f"네이버 금융 page {page} 수집 실패: {e}")

    if all_dfs:
        return pd.concat(all_dfs, ignore_index=True)

    return pd.DataFrame()


def run_analysis():
    st.write("📡 네이버 금융 거래대금 상위 종목 수집 중...")

    df_kospi = get_naver_top_volume(sosok=0, pages=pages)
    df_kospi["구분"] = "KOSPI"

    df_kosdaq = get_naver_top_volume(sosok=1, pages=pages)
    df_kosdaq["구분"] = "KOSDAQ"

    df_naver = pd.concat([df_kospi, df_kosdaq], ignore_index=True)

    if df_naver.empty:
        return pd.DataFrame({"결과": ["네이버 금융 데이터를 가져오지 못했습니다."]})

    st.write(f"원본 수집: {len(df_naver)}개")

    st.write("📡 상장 종목 리스트와 매칭 중...")

    # Streamlit Cloud에서는 FinanceDataReader 상장리스트가 실패할 수 있어서
    # 네이버 데이터 기준으로 ETF/ETN/우선주/스팩을 제외합니다.

    df = df_naver.copy()

    # 시장명 정리
    if "구분" in df.columns:
        df["Market"] = df["구분"]
    else:
        df["Market"] = ""

    # 종목코드는 네이버 거래대금 페이지에는 없으므로 임시로 빈 값 처리
    # 나중에 종목코드가 꼭 필요하면 별도 방식으로 추가 가능
    df["Code"] = ""

    exclude_keywords = [
        "ETF", "ETN", "KODEX", "TIGER", "ACE", "SOL", "KBSTAR",
        "HANARO", "RISE", "KOSEF", "ARIRANG", "TIMEFOLIO",
        "PLUS", "FOCUS", "TREX", "마이티", "히어로즈",
        "레버리지", "인버스", "선물", "TR", "인덱스",
        "스팩", "기업인수목적",
        "우", "우B"
    ]

    def is_excluded_name(name):
        name = str(name)

        for keyword in exclude_keywords:
            if keyword in name:
                return True

        return False

    df = df[~df["종목명"].apply(is_excluded_name)].copy()

    df = df.drop_duplicates(subset=["종목명"]).reset_index(drop=True)

    st.write(f"ETF/ETN/기타 제거 후:run_analysis() {len(df)}개")

    st.write("📊 1차 필터링 중...")

    df["등락률_num"] = df["등락률"].apply(parse_pct)

    df["거래대금"] = pd.to_numeric(df["거래대금"], errors="coerce")
    df["시가총액"] = pd.to_numeric(df["시가총액"], errors="coerce")

    df["거래대금_억"] = df["거래대금"] / 100
    df["시가총액_억"] = df["시가총액"]

    filter1 = df["거래대금_억"] >= min_trade_value

    filter2 = (
        (df["시가총액_억"] >= min_market_cap) &
        (df["시가총액_억"] <= max_market_cap)
    )

    filter3 = (
        (df["등락률_num"] >= min_change) &
        (df["등락률_num"] <= max_change)
    )

    df = df[filter1 & filter2 & filter3].copy()

    if df.empty:
        return pd.DataFrame({"결과": ["조건에 맞는 종목이 없습니다. 조건을 완화해보세요."]})

    df = df.sort_values("등락률", ascending=False)

    # 일봉 데이터로 종가/고가 비율 계산
    st.write("📈 일봉 데이터 분석 중...")

    from datetime import timedelta

    end_date = datetime.now().strftime("%Y-%m-%d")
    start_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")

    ratio_list = []

    for code in df["Code"]:

        try:
            daily = fdr.DataReader(code, start_date, end_date)

            if daily.empty:
                ratio_list.append(None)
            else:
                last = daily.iloc[-1]
                ratio = last["Close"] / last["High"]
                ratio_list.append(round(ratio, 3))

        except:
            ratio_list.append(None)

    df["종가/고가"] = ratio_list

    # 종가/고가 상태 판정
    ratio_status = []

    for r in df["종가/고가"]:

        if r is None:
            ratio_status.append("데이터없음")

        elif 0.88 <= r <= 0.92:
            ratio_status.append("A구간 ⭐")

        elif 0.93 <= r <= 0.95:
            ratio_status.append("관찰 👀")

        else:
            ratio_status.append("제외")

    df["종고비판정"] = ratio_status

    # 3일 패턴 분석: 그제 → 어제 → 오늘
    pattern_list = []
    pattern_meaning_list = []
    pattern_judge_list = []

    def get_trend_symbol(rate):
        if rate >= 1.0:
            return "🔴▲" # 상승
        elif rate <= -1.0:
            return "🔵▼"   # 하락
        else:
            return "🟣→"   # 보합

    def judge_pattern(pattern):
        pattern_map = {
            "🔴▲ 🔴▲ 🔴▲": ("강한 상승 추세 + 매물 소화 후", "⭐ BEST"),
            "🔴▲ 🟣→ 🔴▲": ("상승 후 숨고르기 → 재상승", "✅ OK"),
            "🟣→ 🔴▲ 🔴▲": ("횡보 후 추세 전환", "✅ OK"),
            "🔵▼ 🔴▲ 🔴▲": ("하락 후 반등, 단기 저점 매수", "✅ OK"),
            "🔵▼ 🟣→ 🔴▲": ("반등 신호 약함, 관찰 필요", "⚠️ 관찰"),
            "🔵▼ 🔵▼ 🔴▲": ("V자 반등, 위험", "❌ 위험"),
            "강한 🔴▲ 연속": ("과열", "❌ 제외")
        }

        return pattern_map.get(pattern, ("기타 패턴", "⚠️ 관찰"))

    for code in df["Code"]:
        try:
            daily = fdr.DataReader(code, start_date, end_date)

            if daily.empty or len(daily) < 4:
                pattern_list.append("데이터부족")
                pattern_meaning_list.append("데이터부족")
                pattern_judge_list.append("⚠️ 관찰")
                continue

            closes = daily["Close"].values

            # 그제 등락률
            rate_2days_ago = (closes[-3] - closes[-4]) / closes[-4] * 100

            # 어제 등락률
            rate_yesterday = (closes[-2] - closes[-3]) / closes[-3] * 100

            # 오늘 등락률: 이미 1차 필터에서 +5% 이상이므로 ▲로 처리
            rate_today = (closes[-1] - closes[-2]) / closes[-2] * 100

            s1 = get_trend_symbol(rate_2days_ago)
            s2 = get_trend_symbol(rate_yesterday)
            s3 = get_trend_symbol(rate_today)

            pattern = f"{s1} {s2} {s3}"

            # 강한 상승 연속 과열 조건
            if rate_2days_ago >= 5 and rate_yesterday >= 5:
                pattern = "강한 ▲ 연속"

            meaning, judge = judge_pattern(pattern)

            pattern_list.append(pattern)
            pattern_meaning_list.append(meaning)
            pattern_judge_list.append(judge)

        except:
            pattern_list.append("분석실패")
            pattern_meaning_list.append("분석실패")
            pattern_judge_list.append("⚠️ 관찰")

    df["3일패턴"] = pattern_list
    df["패턴의미"] = pattern_meaning_list
    df["패턴판정"] = pattern_judge_list

    # 10일 폭등이력 분석
    spike_list = []
    spike_judge_list = []

    for code in df["Code"]:
        try:
            daily = fdr.DataReader(code, start_date, end_date)
 
            if daily.empty or len(daily) < 11:
                spike_list.append("데이터부족")
                spike_judge_list.append("⚠️ 관찰")
                continue

            recent_10 = daily.iloc[-11:-1].copy()
            closes = recent_10["Close"].values

            daily_changes = []

            for i in range(1, len(closes)):
                change = (closes[i] - closes[i - 1]) / closes[i - 1] * 100
                daily_changes.append(change)

            spike_count = sum(1 for x in daily_changes if x >= 15)
            cumulative = (closes[-1] - closes[0]) / closes[0] * 100

            spike_text = f"15%↑ {spike_count}회 / 누적 {cumulative:.1f}%"
  
            if spike_count >= 2:
                spike_judge = "❌ 제외"

            elif cumulative >= 30:
                spike_judge = "❌ 제외"
 
            elif spike_count == 1:
                spike_judge = "⚠️ 주의"

            else:
                spike_judge = "✅ 양호"

            spike_list.append(spike_text)
            spike_judge_list.append(spike_judge)

        except:
            spike_list.append("분석실패")
            spike_judge_list.append("⚠️ 관찰")

    df["10일폭등이력"] = spike_list
    df["폭등판정"] = spike_judge_list


    # ============================================================
    # 외국인/기관/개인 수급 분석 - 1거래일 + 최근 3거래일 누적
    # ============================================================

    foreign_1d_list = []
    institution_1d_list = []
    personal_1d_list = []

    foreign_3d_list = []
    institution_3d_list = []
    personal_3d_list = []

    foreign_trend_1d_list = []
    institution_trend_1d_list = []

    foreign_trend_3d_list = []
    institution_trend_3d_list = []

    supply_opinion_list = []
    supply_judge_list = []

    def to_num(v):
        s = str(v)
        s = s.replace(",", "")
        s = s.replace("+", "")
        s = s.strip()

        try:
            return int(float(s))
        except:
            return 0


    def get_supply_data(code):
        url = f"https://finance.naver.com/item/frgn.naver?code={code}"

        try:
            response = requests.get(url, headers=headers, timeout=10)
            response.encoding = "euc-kr"

            tables = pd.read_html(StringIO(response.text))

            for table in tables:
                if table.shape[1] >= 8:
                    temp = table.dropna(how="all").copy()
  
                    try:
                        temp.columns = [
                            "날짜", "종가", "전일비", "등락률", "거래량",
                            "기관", "외국인순매매", "외국인보유", "외국인비율"
                        ]

                        recent = temp.dropna(subset=["날짜"]).head(3)

                        if recent.empty:
                            return None
  
                        recent["기관_n"] = recent["기관"].apply(to_num)
                        recent["외인_n"] = recent["외국인순매매"].apply(to_num)
                        recent["거래량_n"] = recent["거래량"].apply(to_num)

                        recent["기관_n"] = recent["기관"].apply(to_num)
                        recent["외인_n"] = recent["외국인순매매"].apply(to_num)
                        recent["거래량_n"] = recent["거래량"].apply(to_num)

                    # ============================================================
                    # 최근 5일 평균 거래량 계산
                    # ============================================================

                        price_url = f"https://finance.naver.com/item/sise_day.naver?code={code}"

                        price_response = requests.get(
                            price_url,
                            headers=headers,
                            timeout=10
                        )

                        price_response.encoding = "euc-kr"

                        day_df = pd.read_html(
                            StringIO(price_response.text)
                        )[0]

                        day_df = day_df.dropna()

                        day_df["거래량"] = day_df["거래량"].apply(to_num)

                        # 오늘 제외 최근 5일 평균 거래량
                        avg_volume_5d = day_df["거래량"].iloc[1:6].mean()

                        recent["avg_volume_5d"] = avg_volume_5d

                        return recent
                    except:
                        continue

            return None

        except:
            return None


    def trend_text(value, pct):

        if value is None:
            return "—"

        if value > 0:
            return f"🔴매수 {int(value):,}주 ({pct:.1f}%)"

        elif value < 0:
            return f"🔵매도 {int(value):,}주 ({pct:.1f}%)"

        else:
            return "➖보합 0주 (0.0%)"


    for code in df["Code"]:

        recent = get_supply_data(code)

        if recent is None or recent.empty:
            foreign_1d_list.append("—")
            institution_1d_list.append("—")
            personal_1d_list.append("—")

            foreign_3d_list.append("—")
            institution_3d_list.append("—")
            personal_3d_list.append("—")

            foreign_trend_1d_list.append("—")
            institution_trend_1d_list.append("—")
            foreign_trend_3d_list.append("—")
            institution_trend_3d_list.append("—")

            supply_opinion_list.append("수급 데이터 확인 필요")
            supply_judge_list.append("⚠️ 수급확인")
            continue

        # 1거래일
        foreign_1d = recent.iloc[0]["외인_n"]
        institution_1d = recent.iloc[0]["기관_n"]
        volume_1d = recent.iloc[0]["거래량_n"]
        avg_volume_5d = recent["avg_volume_5d"].iloc[0]
        personal_1d = -(foreign_1d + institution_1d)

        # 3거래일 누적
        foreign_3d = recent["외인_n"].sum()
        institution_3d = recent["기관_n"].sum()
        volume_3d = recent["거래량_n"].sum()
        personal_3d = -(foreign_3d + institution_3d)

        foreign_1d_list.append(foreign_1d)
        institution_1d_list.append(institution_1d)
        personal_1d_list.append(personal_1d)

        foreign_3d_list.append(foreign_3d)
        institution_3d_list.append(institution_3d)
        personal_3d_list.append(personal_3d)

        foreign_pct_1d = abs(foreign_1d) / avg_volume_5d * 100 if avg_volume_5d > 0 else 0
        institution_pct_1d = abs(institution_1d) / avg_volume_5d * 100 if avg_volume_5d > 0 else 0

        foreign_pct_3d = abs(foreign_3d) / (avg_volume_5d * 3) * 100 if avg_volume_5d > 0 else 0
        institution_pct_3d = abs(institution_3d) / (avg_volume_5d * 3) * 100 if avg_volume_5d > 0 else 0

        foreign_trend_1d_list.append(trend_text(foreign_1d, foreign_pct_1d))
        institution_trend_1d_list.append(trend_text(institution_1d, institution_pct_1d))

        foreign_trend_3d_list.append(trend_text(foreign_3d, foreign_pct_3d))
        institution_trend_3d_list.append(trend_text(institution_3d, institution_pct_3d))

        # 판정은 3거래일 누적 기준
        foreign_pct_3d = abs(foreign_3d) / volume_3d * 100 if volume_3d > 0 else 0
        institution_pct_3d = abs(institution_3d) / volume_3d * 100 if volume_3d > 0 else 0
        total_sell_pct = foreign_pct_3d + institution_pct_3d

        if foreign_3d > 0 or institution_3d > 0:
            opinion = "✅ 양호: 최근 3일 외인 또는 기관 순매수"
            judge = "✅ 양호"

        elif foreign_3d < 0 and institution_3d < 0 and personal_3d > 0:
            if total_sell_pct >= 3:
                opinion = "❌ 제외: 최근 3일 외인+기관 동시 매도 / 개인 매수"
                judge = "❌ 수급제외"
            else:
                opinion = "⚠️ 주의: 최근 3일 외인+기관 매도지만 비중 약함"
                judge = "⚠️ 주의"

        else:
            opinion = "➖ 중립: 뚜렷한 수급 방향 없음"
            judge = "➖ 중립"

        supply_opinion_list.append(opinion)
        supply_judge_list.append(judge)


    df["외인_1일"] = foreign_1d_list
    df["기관_1일"] = institution_1d_list
    df["개인추정_1일"] = personal_1d_list

    df["외인추세_1일"] = foreign_trend_1d_list
    df["기관추세_1일"] = institution_trend_1d_list

    df["외인_3일"] = foreign_3d_list
    df["기관_3일"] = institution_3d_list
    df["개인추정_3일"] = personal_3d_list

    df["외인추세_3일"] = foreign_trend_3d_list
    df["기관추세_3일"] = institution_trend_3d_list

    df["수급의견"] = supply_opinion_list
    df["수급판정"] = supply_judge_list

    result = df[[
        "Code",
        "종목명",
        "Market",
        "현재가",
        "등락률",
        "종가/고가",
        "종고비판정",
        "3일패턴",
        "패턴의미",
        "패턴판정",
        "10일폭등이력",
        "폭등판정",
        "외인추세_1일",
        "기관추세_1일",
        "개인추정_1일",
        "외인추세_3일",
        "기관추세_3일",
        "개인추정_3일",
        "수급의견",
        "수급판정",
        "거래량",
        "거래대금_억",
        "시가총액_억"
    ]].copy()

    result.columns = [
        "종목코드",
        "종목명",
        "시장",
        "현재가",
        "등락률",
        "종가/고가",
        "종고비판정",
        "3일패턴",
        "패턴의미",
        "패턴판정",
        "10일폭등이력",
        "폭등판정",
        "외인추세_1일",
        "기관추세_1일",
        "개인추정_1일",
        "외인추세_3일",
        "기관추세_3일",
        "개인추정_3일",
        "수급의견",
        "수급판정",
        "거래량",
        "거래대금(억)",
        "시총(억)"
    ]

    # 숫자 쉼표 추가
    result["현재가"] = result["현재가"].apply(
    lambda x: f"{int(x):,}"
    )

    result["거래량"] = result["거래량"].apply(
    lambda x: f"{int(x):,}"
    )

    result["거래대금(억)"] = result["거래대금(억)"].apply(
    lambda x: f"{float(x):,.2f}"
    )

    result["시총(억)"] = result["시총(억)"].apply(
    lambda x: f"{int(x):,}"
    )

    st.success(f"1차 필터 통과: {len(result)}개 종목")

    return result


def make_excel(df):
    output = BytesIO()

    wb = Workbook()
    ws = wb.active
    ws.title = "매매후보"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    normal_font = Font(name="맑은 고딕", size=10)
    red_font = Font(name="맑은 고딕", size=10, bold=True, color="D32F2F")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.alignment = center
            cell.border = border

            col_name = df.columns[col_idx - 1]

            if col_name in ["등락률"]:
                cell.font = red_font

    widths = {
        "종목코드": 12,
        "종목명": 18,
        "시장": 10,
        "현재가": 12,
        "등락률": 12,
        "거래량": 15,
        "거래대금(억)": 15,
        "시총(억)": 15,
        "결과": 45
    }

    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = widths.get(col_name, 15)

    ws.freeze_panes = "A2"
    wb.save(output)
    output.seek(0)

    return output

if "df_result" not in st.session_state:
    st.session_state.df_result = None

if st.button("🔍 오늘 단타 후보 분석 시작"):
    with st.spinner("분석 중입니다..."):
        st.session_state.df_result = run_analysis()

if st.session_state.df_result is not None:

    st.success("분석 완료!")

    st.subheader("📊 분석 결과")

    view_df = st.session_state.df_result.copy()

    if "숨김" not in view_df.columns:
        view_df.insert(0, "숨김", False)

    edited_df = st.data_editor(
        view_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "숨김": st.column_config.CheckboxColumn(
                "숨김",
                help="체크하면 아래 최종 결과에서 제외됩니다.",
                default=False,
            )
        },
        disabled=[col for col in view_df.columns if col != "숨김"]
    )

    final_view = edited_df[
        edited_df["숨김"] == False
    ].drop(columns=["숨김"])

    st.subheader("✅ 최종 확인 종목")
    st.dataframe(final_view, use_container_width=True)

    excel_file = make_excel(final_view)

    today = datetime.now().strftime("%Y-%m-%d")

    st.download_button(
        label="📥 엑셀 다운로드",
        data=excel_file,
        file_name=f"단타분석_{today}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )